"""Excel-шаблоны: валидация при загрузке и генерация (ADR-0005 §Excel)."""

from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from re import Match

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from domain.template_markers import (
    CONTRACT_VERSION,
    MARKER_RE,
    MarkerSyntaxError,
    RowBlockSpec,
    block_name_from_token,
    canonical_key,
    extract_markers,
    find_malformed_marker_fragments,
    is_structural_token,
    validate_block_name,
)


class ExcelTemplateError(Exception):
    """Ошибка Excel-шаблона."""


class TemplateValidationError(ExcelTemplateError):
    def __init__(self, message: str, *, unknown_markers: tuple[str, ...] = ()) -> None:
        super().__init__(message)
        self.unknown_markers = unknown_markers


@dataclass(frozen=True)
class ArchivedTemplate:
    archive_path: Path
    contract_version: str = CONTRACT_VERSION


def archive_upload(source: Path, archive_path: Path) -> ArchivedTemplate:
    """Сохранить исходник byte-for-byte и проверить маркеры (ADR-0005)."""
    validate_archived(source)
    data = source.read_bytes()
    archive_path.parent.mkdir(parents=True, exist_ok=True)
    archive_path.write_bytes(data)
    return ArchivedTemplate(archive_path=archive_path)


def validate_archived(path: Path) -> None:
    book = load_workbook(path, data_only=False)
    try:
        _validate_workbook(book)
    finally:
        book.close()


def generate_excel_report(
    archived: ArchivedTemplate,
    output_path: Path,
    *,
    scalars: dict[str, str],
    row_records: list[dict[str, str]] | None = None,
    named_row_records: dict[str, list[dict[str, str]]] | None = None,
) -> None:
    """Производный файл; архивный исходник не изменяется."""
    shutil.copyfile(archived.archive_path, output_path)
    book = load_workbook(output_path)
    try:
        records = row_records or []
        named = named_row_records or {}
        for sheet in book.worksheets:
            blocks = _blocks_for_sheet(sheet)
            block_rows = _rows_in_blocks(blocks)
            _substitute_scalars(sheet, scalars, skip_rows=block_rows)
            for block in sorted(blocks, key=lambda b: b.start_row, reverse=True):
                block_records = records if block.name is None else named.get(block.name, [])
                _expand_block(sheet, block, block_records)
        book.save(output_path)
    finally:
        book.close()


def _validate_workbook(book: Workbook) -> None:
    unknown: set[str] = set()
    for sheet in book.worksheets:
        unknown |= _validate_sheet(sheet)
    if unknown:
        names = ", ".join(f"{{{{{m}}}}}" for m in sorted(unknown))
        raise TemplateValidationError(
            f"unknown markers: {names}",
            unknown_markers=tuple(sorted(unknown)),
        )


def _validate_sheet(sheet: Worksheet) -> set[str]:
    unknown: set[str] = set()
    unnamed_starts: list[int] = []
    unnamed_ends: list[int] = []
    named_starts: dict[str, int] = {}
    named_ends: dict[str, int] = {}

    for row in sheet.iter_rows():
        for cell in row:
            text = _cell_source_text(cell)
            if not text:
                continue
            malformed = find_malformed_marker_fragments(text)
            if malformed:
                raise TemplateValidationError(
                    "malformed marker syntax in sheet "
                    f"{sheet.title!r} cell {cell.coordinate}: {malformed[0]!r}"
                )
            try:
                tokens = extract_markers(text)
            except MarkerSyntaxError as exc:
                raise TemplateValidationError(str(exc)) from exc
            for token in tokens:
                if token == "#ROW":
                    unnamed_starts.append(cell.row)
                elif token == "/ROW":
                    unnamed_ends.append(cell.row)
                elif token.startswith("#ROW:"):
                    name = block_name_from_token(token)
                    if name is None or not validate_block_name(name):
                        raise TemplateValidationError(f"invalid block name in {{{{{token}}}}}")
                    named_starts[name] = cell.row
                elif token.startswith("/ROW:"):
                    name = block_name_from_token(token)
                    if name is None or not validate_block_name(name):
                        raise TemplateValidationError(f"invalid block name in {{{{{token}}}}}")
                    named_ends[name] = cell.row
                elif is_structural_token(token):
                    raise TemplateValidationError(f"unknown structural marker: {{{{{token}}}}}")
                elif canonical_key(token) is None:
                    unknown.add(token)

    if len(unnamed_starts) > 1:
        raise TemplateValidationError("at most one unnamed {{#ROW}} block per sheet")
    if unnamed_ends and not unnamed_starts:
        raise TemplateValidationError("{{/ROW}} without {{#ROW}}")
    for name, start in named_starts.items():
        if name not in named_ends:
            raise TemplateValidationError(f"unclosed block {{#ROW:{name}}}")
        if named_ends[name] < start:
            raise TemplateValidationError(f"block {{#ROW:{name}}} ends before it starts")
    for name in named_ends:
        if name not in named_starts:
            raise TemplateValidationError(f"unopened block {{/ROW:{name}}}")

    return unknown


def _blocks_for_sheet(sheet: Worksheet) -> list[RowBlockSpec]:
    unnamed_start: int | None = None
    unnamed_end: int | None = None
    named_starts: dict[str, int] = {}
    named_ends: dict[str, int] = {}

    for row in sheet.iter_rows():
        for cell in row:
            for token in extract_markers(_cell_source_text(cell)):
                if token == "#ROW":
                    unnamed_start = cell.row
                elif token == "/ROW":
                    unnamed_end = cell.row
                elif token.startswith("#ROW:"):
                    name = block_name_from_token(token)
                    if name is not None:
                        named_starts[name] = cell.row
                elif token.startswith("/ROW:"):
                    name = block_name_from_token(token)
                    if name is not None:
                        named_ends[name] = cell.row

    blocks: list[RowBlockSpec] = []
    if unnamed_start is not None:
        end = unnamed_end if unnamed_end is not None else unnamed_start
        blocks.append(RowBlockSpec(name=None, start_row=unnamed_start, end_row=end))
    for name, start in named_starts.items():
        blocks.append(RowBlockSpec(name=name, start_row=start, end_row=named_ends[name]))
    return blocks


def _rows_in_blocks(blocks: list[RowBlockSpec]) -> set[int]:
    rows: set[int] = set()
    for block in blocks:
        rows.update(range(block.start_row, block.end_row + 1))
    return rows


def _substitute_scalars(sheet: Worksheet, scalars: dict[str, str], *, skip_rows: set[int]) -> None:
    for row in sheet.iter_rows():
        if row[0].row in skip_rows:
            continue
        for cell in row:
            text = _cell_source_text(cell)
            if text and "{{" in text:
                cell.value = _substitute_text(text, scalars)


def _expand_block(sheet: Worksheet, block: RowBlockSpec, records: list[dict[str, str]]) -> None:
    height = block.end_row - block.start_row + 1
    snapshot = _snapshot_rows(sheet, block.start_row, block.end_row)
    if not records:
        records = [{}]
    _apply_block(sheet, block.start_row, snapshot, records[0], template_row=block.start_row)
    insert_at = block.end_row + 1
    for record in records[1:]:
        sheet.insert_rows(insert_at, height)
        _paste_block(sheet, insert_at, snapshot, template_row=block.start_row)
        _apply_block(sheet, insert_at, snapshot, record, template_row=block.start_row)
        insert_at += height


def _snapshot_rows(sheet: Worksheet, start: int, end: int) -> list[list[object | None]]:
    rows: list[list[object | None]] = []
    for r in range(start, end + 1):
        row_vals: list[object | None] = []
        for col in range(1, sheet.max_column + 1):
            row_vals.append(sheet.cell(r, col).value)
        rows.append(row_vals)
    return rows


def _paste_block(
    sheet: Worksheet,
    dest_start: int,
    snapshot: list[list[object | None]],
    *,
    template_row: int,
) -> None:
    for offset, row_vals in enumerate(snapshot):
        dest_row = dest_start + offset
        src_row = template_row + offset
        for col, value in enumerate(row_vals, start=1):
            src = sheet.cell(src_row, col)
            dst = sheet.cell(dest_row, col)
            if isinstance(value, str) and value.startswith("="):
                origin = f"{get_column_letter(col)}{src_row}"
                target = f"{get_column_letter(col)}{dest_row}"
                dst.value = Translator(value, origin=origin).translate_formula(target)
            else:
                dst.value = value
            _copy_style(src, dst)


def _apply_block(
    sheet: Worksheet,
    start: int,
    snapshot: list[list[object | None]],
    record: dict[str, str],
    *,
    template_row: int,
) -> None:
    for offset in range(len(snapshot)):
        row_idx = start + offset
        for col in range(1, len(snapshot[offset]) + 1):
            cell = sheet.cell(row_idx, col)
            text = _cell_source_text(cell)
            if text and "{{" in text:
                cell.value = _substitute_text(text, record)


def _substitute_text(text: str, values: dict[str, str]) -> str:
    def repl(match: Match[str]) -> str:
        token = match.group(1)
        if is_structural_token(token):
            return ""
        key = canonical_key(token)
        if key is None:
            return match.group(0)
        return values.get(key, "")

    return MARKER_RE.sub(repl, text)


def _cell_source_text(cell: object) -> str:
    value = getattr(cell, "value", None)
    if value is None:
        return ""
    return str(value)


def _copy_style(src: object, dst: object) -> None:
    for attr in ("font", "fill", "border", "alignment", "number_format", "protection"):
        if hasattr(src, attr):
            setattr(dst, attr, copy(getattr(src, attr)))
