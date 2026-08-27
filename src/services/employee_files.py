"""Чтение/запись сырых таблиц сотрудников (CSV и XLSX)."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from domain.employee_import import cell_text


class EmployeeFileError(ValueError):
    """Некорректный или нечитаемый файл импорта/экспорта."""


def read_tabular(path: Path) -> tuple[list[str], list[list[str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise EmployeeFileError(f"unsupported format: {suffix or path.name}")


def write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    book.save(path)


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = [cell_text(c) for c in next(reader)]
        except StopIteration as exc:
            raise EmployeeFileError("empty csv") from exc
        rows = [[cell_text(c) for c in row] for row in reader]
    return headers, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book.active
        if sheet is None:
            raise EmployeeFileError("empty workbook")
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise EmployeeFileError("empty workbook") from exc
        headers = [cell_text(c) for c in header_row]
        rows = [[cell_text(c) for c in row] for row in iterator]
    finally:
        book.close()
    return headers, rows
