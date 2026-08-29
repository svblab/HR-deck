"""Unit/integration: Excel template engine (ADR-0005 §Excel)."""

from __future__ import annotations

import hashlib
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from domain.template_markers import canonical_key
from reports.excel_template import (
    TemplateValidationError,
    archive_upload,
    generate_excel_report,
    validate_archived,
)


def _save_template(path: Path, rows: list[list[str | None]]) -> None:
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            if value is not None:
                sheet.cell(r, c, value)
    book.save(path)
    book.close()


def test_adr0005_excel_unknown_marker_rejected_on_upload(tmp_path: Path) -> None:
    src = tmp_path / "bad.xlsx"
    _save_template(src, [["{{not_in_catalog}}", "{{ФИО}}"]])
    archive = tmp_path / "archive.xlsx"
    with pytest.raises(TemplateValidationError) as exc:
        archive_upload(src, archive)
    assert "not_in_catalog" in exc.value.unknown_markers
    assert not archive.exists()


def test_adr0005_excel_alias_resolves_like_canonical_key() -> None:
    assert canonical_key("ФИО") == "employee.full_name"
    assert canonical_key("employee.full_name") == "employee.full_name"


def test_adr0005_excel_two_unnamed_row_blocks_rejected(tmp_path: Path) -> None:
    src = tmp_path / "two_blocks.xlsx"
    _save_template(
        src,
        [
            ["{{#ROW}}", "{{ФИО}}"],
            ["{{#ROW}}", "{{должность}}"],
        ],
    )
    with pytest.raises(TemplateValidationError, match="at most one unnamed"):
        validate_archived(src)


def test_adr0005_excel_named_block_mismatch_rejected(tmp_path: Path) -> None:
    src = tmp_path / "mismatch.xlsx"
    _save_template(
        src,
        [
            ["{{#ROW:alpha}}", "{{ФИО}}"],
            ["{{/ROW:beta}}", ""],
        ],
    )
    with pytest.raises(TemplateValidationError, match="unclosed block"):
        validate_archived(src)


def test_adr0005_excel_row_block_expands_preserving_format(tmp_path: Path) -> None:
    src = tmp_path / "styled.xlsx"
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet["A1"] = "Report"
    sheet["A2"] = "{{#ROW}}"
    cell = sheet["B2"]
    cell.value = "{{ФИО}}"
    cell.font = Font(bold=True, color="FF0000")
    cell.fill = PatternFill("solid", fgColor="FFFF00")
    book.save(src)
    book.close()

    archive = tmp_path / "store" / "orig.xlsx"
    archived = archive_upload(src, archive)
    out = tmp_path / "out.xlsx"
    generate_excel_report(
        archived,
        out,
        scalars={},
        row_records=[
            {"employee.full_name": "Иванов"},
            {"employee.full_name": "Петров"},
        ],
    )
    from openpyxl import load_workbook

    result = load_workbook(out)
    sheet = result.active
    assert sheet is not None
    assert sheet["B2"].value == "Иванов"
    assert sheet["B3"].value == "Петров"
    assert sheet["B2"].font.bold
    assert sheet["B2"].font.color and sheet["B2"].font.color.rgb == "00FF0000"
    assert sheet["B3"].font.bold
    assert sheet["B3"].fill.fgColor.rgb == "00FFFF00"
    result.close()


def test_adr0005_excel_missing_value_renders_empty_string(tmp_path: Path) -> None:
    src = tmp_path / "row.xlsx"
    _save_template(src, [["{{#ROW}}", "{{ФИО}}", "{{должность}}"]])
    archive = tmp_path / "orig.xlsx"
    archived = archive_upload(src, archive)
    out = tmp_path / "out.xlsx"
    generate_excel_report(archived, out, scalars={}, row_records=[{"employee.full_name": "X"}])
    from openpyxl import load_workbook

    book = load_workbook(out)
    sheet = book.active
    assert sheet is not None
    assert sheet["B1"].value == "X"
    assert sheet["C1"].value in ("", None)
    book.close()


def test_adr0005_excel_archive_unchanged_after_generation(tmp_path: Path) -> None:
    src = tmp_path / "tpl.xlsx"
    _save_template(src, [["Title {{заголовок}}"], ["{{#ROW}}", "{{ФИО}}"]])
    archive = tmp_path / "orig.xlsx"
    archived = archive_upload(src, archive)
    before = hashlib.sha256(archive.read_bytes()).hexdigest()
    generate_excel_report(
        archived,
        tmp_path / "out.xlsx",
        scalars={"report.title": "T"},
        row_records=[{"employee.full_name": "A"}, {"employee.full_name": "B"}],
    )
    after = hashlib.sha256(archive.read_bytes()).hexdigest()
    assert before == after


def test_adr0005_excel_scalar_outside_row_block(tmp_path: Path) -> None:
    src = tmp_path / "scalar.xlsx"
    _save_template(src, [["Period: {{период_с}} — {{период_по}}"], ["{{#ROW}}", "{{ФИО}}"]])
    archive = tmp_path / "orig.xlsx"
    archived = archive_upload(src, archive)
    out = tmp_path / "out.xlsx"
    generate_excel_report(
        archived,
        out,
        scalars={"report.period_from": "2026-01-01", "report.period_to": "2026-01-31"},
        row_records=[{"employee.full_name": "One"}],
    )
    from openpyxl import load_workbook

    book = load_workbook(out)
    sheet = book.active
    assert sheet is not None
    assert sheet["A1"].value == "Period: 2026-01-01 — 2026-01-31"
    assert sheet["B2"].value == "One"
    book.close()


def test_adr0005_excel_formula_outside_block_survives_row_insert(tmp_path: Path) -> None:
    src = tmp_path / "formula.xlsx"
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet["A2"] = "{{#ROW}}"
    sheet["B2"] = "{{ФИО}}"
    sheet["A4"] = "=B2"
    book.save(src)
    book.close()

    archive = tmp_path / "orig.xlsx"
    archived = archive_upload(src, archive)
    out = tmp_path / "out.xlsx"
    generate_excel_report(
        archived,
        out,
        scalars={},
        row_records=[
            {"employee.full_name": "First"},
            {"employee.full_name": "Second"},
        ],
    )
    from openpyxl import load_workbook

    book = load_workbook(out)
    sheet = book.active
    assert sheet is not None
    assert sheet["B2"].value == "First"
    assert sheet["B3"].value == "Second"
    # Formula row shifted down by one inserted row; still references first data cell.
    assert sheet["A5"].value == "=B2"
    book.close()
