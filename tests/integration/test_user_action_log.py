"""Integration: реальные действия видны в журнале; Excel — только отфильтрованные строки."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from domain.action_log import EXPORT_HEADERS, ActionLogFilters
from domain.employee import EmployeeCreateInput
from domain.permissions import RoleCode
from services.account_management import AccountManagementService
from services.bootstrap import BootstrapService
from services.employees import EmployeeService
from services.user_action_log import UserActionLogService
from tests.fixtures.synthetic import seed_synthetic_org


def _open(tmp_path: Path):
    clock = lambda: "2026-08-26T14:00:00Z"  # noqa: E731
    conn, session, _code = BootstrapService(clock=clock).initial_administrator_setup(
        db_path=tmp_path / "app.db", login="admin", password="AdminPass-1"
    )
    ids = seed_synthetic_org(conn)
    employees = EmployeeService(conn, session, clock=clock)
    accounts = AccountManagementService(
        conn, session, db_path=tmp_path / "app.db", clock=clock
    )
    log = UserActionLogService(conn, session)
    return conn, session, employees, accounts, log, ids


@pytest.mark.acceptance
def test_journal_surfaces_create_employee_and_reset_password(tmp_path: Path) -> None:
    conn, _session, employees, accounts, log, ids = _open(tmp_path)
    emp_id = employees.create_employee(
        EmployeeCreateInput(
            full_name="Петров Пётр",
            position_id=ids["position_engineer_id"],
            branch_id=ids["branch_id"],
            department_id=ids["department_id"],
            division_id=ids["division_id"],
            employment_type_id=1,
        )
    )
    hr_id = accounts.create_account(login="hr1", password="HrPass-1", role=RoleCode.HR_EMPLOYEE)
    accounts.reset_password(hr_id, "HrPass-2")

    created = log.list_entries(ActionLogFilters(action_type="employee.create", employee_id=emp_id))
    assert len(created) == 1
    assert created[0].entity_type == "employee"
    assert created[0].entity_id == emp_id
    assert created[0].result == "success"

    resets = log.list_entries(ActionLogFilters(action_type="account.reset_password"))
    assert [r.entity_id for r in resets] == [hr_id]

    by_employee = log.list_entries(ActionLogFilters(employee_id=emp_id))
    assert {r.action_type for r in by_employee} == {"employee.create"}

    before = len(log.list_entries())
    log.list_entries()
    assert len(log.list_entries()) == before
    conn.close()


@pytest.mark.acceptance
def test_filtered_xlsx_contains_only_matching_rows(tmp_path: Path) -> None:
    conn, _session, employees, accounts, log, ids = _open(tmp_path)
    employees.create_employee(
        EmployeeCreateInput(
            full_name="Сидоров Сидор",
            position_id=ids["position_engineer_id"],
            branch_id=ids["branch_id"],
            department_id=ids["department_id"],
            employment_type_id=1,
        )
    )
    accounts.create_account(login="hr1", password="HrPass-1", role=RoleCode.HR_EMPLOYEE)
    path = tmp_path / "journal.xlsx"
    filters = ActionLogFilters(action_type="employee.create")
    count = log.export_xlsx(path, filters)
    expected = log.list_entries(filters)
    assert count == len(expected)
    book = load_workbook(path)
    sheet = book.active
    assert [cell.value for cell in sheet[1]] == list(EXPORT_HEADERS)
    body = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(body) == count
    assert all(row[2] == "employee.create" for row in body)
    conn.close()
